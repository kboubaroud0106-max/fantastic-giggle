from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, make_response
import os
import io
import csv
import json
import base64
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch

from database import (
    init_db, save_response, get_all_responses, get_dashboard_stats,
    verify_admin_credentials, update_admin_password, create_manual_backup,
    list_backups, restore_backup, delete_response, get_db_connection
)
from stats import compile_full_stats

app = Flask(__name__)
app.secret_key = "safi_cinema_survey_secret_key_2026_!@#"
app.config['UPLOAD_FOLDER'] = 'backups'

# Ensure DB initialized
init_db()

# Decorator to protect admin routes
def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def home():
    return render_template('survey.html')

@app.route('/submit', methods=['POST'])
def submit_survey():
    try:
        data = request.json
        if not data:
            return jsonify({"status": "error", "message": "Aucune donnée reçue"}), 400
            
        # Basic validation
        if not data.get('q1_gender') or not data.get('q2_age_group') or not data.get('q5_education_level'):
            return jsonify({"status": "error", "message": "Veuillez remplir les questions obligatoires du profil."}), 400
            
        response_id = save_response(data)
        return jsonify({"status": "success", "message": "Merci ! Votre participation a été enregistrée avec succès.", "id": response_id})
    except Exception as e:
        print(f"Error in submit: {e}")
        return jsonify({"status": "error", "message": "Une erreur interne est survenue lors de l'enregistrement."}), 500

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('logged_in'):
        return redirect(url_for('admin_dashboard'))
        
    error = None
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if verify_admin_credentials(username, password):
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('admin_dashboard'))
        else:
            error = "Identifiant ou mot de passe incorrect."
            
    return render_template('login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    stats = get_dashboard_stats()
    return render_template('admin.html', stats=stats)

@app.route('/admin/api/stats')
@login_required
def get_stats_api():
    filters = {
        'gender': request.args.get('gender'),
        'age_group': request.args.get('age_group'),
        'neighborhood': request.args.get('neighborhood'),
        'education_level': request.args.get('education_level'),
        'profession': request.args.get('profession')
    }
    # Clean filters (remove empty values)
    filters = {k: v for k, v in filters.items() if v}
    
    responses = get_all_responses(filters)
    compiled = compile_full_stats(responses)
    return jsonify(compiled)

@app.route('/admin/api/responses')
@login_required
def get_responses_api():
    filters = {
        'gender': request.args.get('gender'),
        'age_group': request.args.get('age_group'),
        'neighborhood': request.args.get('neighborhood'),
        'education_level': request.args.get('education_level'),
        'profession': request.args.get('profession')
    }
    filters = {k: v for k, v in filters.items() if v}
    
    keyword = request.args.get('search', '').strip().lower()
    
    responses = get_all_responses(filters)
    
    # Keyword search filtering (across all text fields)
    if keyword:
        filtered_responses = []
        for r in responses:
            found = False
            # Check open response text fields
            text_to_search = [
                r.get('q3_neighborhood') or '',
                r.get('q12_memory') or '',
                r.get('q13_text') or '',
                r.get('q16_main_cause') or '',
                r.get('q18_meaning') or '',
                r.get('q26_comments') or '',
                r.get('q27_contact_details') or ''
            ]
            # Search inside Q13 cinema list
            for c in r.get('q13_table', []):
                text_to_search.append(c.get('name') or '')
                text_to_search.append(c.get('location') or '')
                text_to_search.append(c.get('current_state') or '')
                
            for text in text_to_search:
                if keyword in text.lower():
                    found = True
                    break
            if found:
                filtered_responses.append(r)
        responses = filtered_responses
        
    return jsonify(responses)

@app.route('/admin/api/responses/<int:response_id>', methods=['GET'])
@login_required
def get_single_response(response_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM responses WHERE id = ?", (response_id,))
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return jsonify({"status": "error", "message": "Réponse introuvable"}), 404
        
    row_dict = dict(row)
    # Parse JSON fields
    for field in ['q8_periods', 'q10_companions', 'q11_movie_types', 'q14_what_became', 'q20_desired_usage', 'q21_support_type', 'q23_channels']:
        if row_dict.get(field):
            try:
                row_dict[field] = json.loads(row_dict[field])
            except Exception:
                row_dict[field] = []
        else:
            row_dict[field] = []
            
    cursor.execute("SELECT name, location, current_state FROM cinema_mentions WHERE response_id = ?", (response_id,))
    row_dict['q13_table'] = [dict(c) for c in cursor.fetchall()]
    
    conn.close()
    return jsonify(row_dict)

@app.route('/admin/api/responses/<int:response_id>/delete', methods=['POST'])
@login_required
def delete_single_response(response_id):
    try:
        delete_response(response_id)
        return jsonify({"status": "success", "message": f"Réponse #{response_id} supprimée avec succès."})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Erreur lors de la suppression : {e}"}), 500

# Security: Password Update
@app.route('/admin/api/update_password', methods=['POST'])
@login_required
def change_password():
    data = request.json
    old_pass = data.get('old_password')
    new_pass = data.get('new_password')
    
    if not old_pass or not new_pass:
        return jsonify({"status": "error", "message": "Veuillez remplir tous les champs."}), 400
        
    if verify_admin_credentials(session['username'], old_pass):
        update_admin_password(session['username'], new_pass)
        return jsonify({"status": "success", "message": "Mot de passe modifié avec succès."})
    else:
        return jsonify({"status": "error", "message": "Ancien mot de passe incorrect."}), 400

# Backups Management
@app.route('/admin/api/backup', methods=['POST'])
@login_required
def run_backup():
    filename = create_manual_backup()
    if filename:
        return jsonify({"status": "success", "message": f"Sauvegarde créée : {filename}", "backups": list_backups()})
    return jsonify({"status": "error", "message": "Erreur lors de la sauvegarde."}), 500

@app.route('/admin/api/backup/list', methods=['GET'])
@login_required
def get_backups_list():
    return jsonify(list_backups())

@app.route('/admin/api/backup/restore', methods=['POST'])
@login_required
def do_restore():
    data = request.json
    filename = data.get('filename')
    if not filename:
        return jsonify({"status": "error", "message": "Aucun fichier sélectionné"}), 400
        
    if restore_backup(filename):
        return jsonify({"status": "success", "message": "Base de données restaurée avec succès."})
    return jsonify({"status": "error", "message": "La restauration a échoué. Retour à l'état précédent."}), 500

@app.route('/admin/api/backup/upload', methods=['POST'])
@login_required
def upload_backup_file():
    if 'backup_file' not in request.files:
        return jsonify({"status": "error", "message": "Aucun fichier fourni"}), 400
        
    file = request.files['backup_file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Nom de fichier vide"}), 400
        
    if file and file.filename.endswith('.db'):
        filename = f"uploaded_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        file.save(filepath)
        
        # Now restore it
        if restore_backup(filename):
            return jsonify({"status": "success", "message": "Sauvegarde importée et restaurée avec succès."})
        else:
            return jsonify({"status": "error", "message": "Échec de la restauration à partir du fichier importé."}), 500
            
    return jsonify({"status": "error", "message": "Le fichier doit être une base SQLite avec l'extension .db"}), 400

# Exports: CSV & Excel
@app.route('/admin/api/export/csv')
@login_required
def export_csv():
    filters = {
        'gender': request.args.get('gender'),
        'age_group': request.args.get('age_group'),
        'neighborhood': request.args.get('neighborhood'),
        'education_level': request.args.get('education_level'),
        'profession': request.args.get('profession')
    }
    filters = {k: v for k, v in filters.items() if v}
    responses = get_all_responses(filters)
    
    si = io.StringIO()
    cw = csv.writer(si, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    # Headers
    headers = [
        "N Questionnaire", "Date de soumission", "Q1_Genre", "Q2_Age", "Q3_Quartier", "Q4_Duree_Safi", 
        "Q5_Niveau_Etudes", "Q6_Profession", "Q7_Frequente_Cinema", "Q8_Periodes", "Q9_Frequence", 
        "Q10_Accompagnateurs", "Q11_Types_Films", "Q11_Autres_Films", "Q12_Souvenir_Marquant", "Q13_Texte", 
        "Q14_Devenir_Salles", "Q15_1_TV", "Q15_2_Internet", "Q15_3_Piratage", "Q15_4_Prix_Billets", 
        "Q15_5_Vetuste", "Q15_6_Qualite_Films", "Q15_7_Insecurite", "Q15_8_Valeur_Terrains", 
        "Q15_9_Manque_Soutien", "Q15_10_Habitudes_Loisirs", "Q16_Cause_Principale", "Q17_1_Souvenirs", 
        "Q17_2_Perte_Safi", "Q17_3_Etat_Degrade", "Q17_4_Attachement", "Q17_5_Identite_Safi", 
        "Q17_6_Jeunes_Ignorent", "Q18_Representation_Un_Mot", "Q19_1_Patrimoine", "Q19_2_Preserver", 
        "Q19_3_Dynamiser_Centre", "Q19_4_Creer_Emplois", "Q19_5_Frequentation_Futur", "Q19_6_Associer_Habitants", 
        "Q20_Usage_Souhaite", "Q20_Autre_Usage", "Q21_Soutenir_Projet", "Q22_Vu_Lu_Contenu", 
        "Q23_Canaux", "Q24_Suivre_Pages", "Q25_1_Pas_Assez_Medias", "Q25_2_Aide_Reseaux", "Q25_3_Savoir_Plus", 
        "Q26_Commentaires", "Q27_Recontacte", "Q27_Moyen_Contact"
    ]
    cw.writerow(headers)
    
    for r in responses:
        cw.writerow([
            r['id'], r['submission_date'], r['q1_gender'], r['q2_age_group'], r['q3_neighborhood'], r['q4_residence_duration'],
            r['q5_education_level'], r['q6_profession'], r['q7_visited_cinema'], ", ".join(r['q8_periods']), r['q9_frequency'],
            ", ".join(r['q10_companions']), ", ".join(r['q11_movie_types']), r['q11_other'], r['q12_memory'], r['q13_text'],
            ", ".join(r['q14_what_became']), r['q15_1'], r['q15_2'], r['q15_3'], r['q15_4'],
            r['q15_5'], r['q15_6'], r['q15_7'], r['q15_8'], r['q15_9'], r['q15_10'], r['q16_main_cause'],
            r['q17_1'], r['q17_2'], r['q17_3'], r['q17_4'], r['q17_5'], r['q17_6'], r['q18_meaning'],
            r['q19_1'], r['q19_2'], r['q19_3'], r['q19_4'], r['q19_5'], r['q19_6'],
            ", ".join(r['q20_desired_usage']), r['q20_other'], ", ".join(r['q21_support_type']), r['q22_seen_content'],
            ", ".join(r['q23_channels']), r['q24_follow_pages'], r['q25_1'], r['q25_2'], r['q25_3'],
            r['q26_comments'], r['q27_recontact'], r['q27_contact_details']
        ])
        
    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename=enquete_safi_cinema_{datetime.now().strftime('%Y%m%d')}.csv"
    output.headers["Content-type"] = "text/csv; charset=utf-8-sig"
    return output

@app.route('/admin/api/export/excel')
@login_required
def export_excel():
    filters = {
        'gender': request.args.get('gender'),
        'age_group': request.args.get('age_group'),
        'neighborhood': request.args.get('neighborhood'),
        'education_level': request.args.get('education_level'),
        'profession': request.args.get('profession')
    }
    filters = {k: v for k, v in filters.items() if v}
    responses = get_all_responses(filters)
    
    wb = Workbook()
    
    # Sheet 1: Responses details
    ws1 = wb.active
    ws1.title = "Reponses Enquete"
    
    # Design colors
    title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    text_font = Font(name="Calibri", size=11)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    headers = [
        "N° Questionnaire", "Date de soumission", "Q1. Genre", "Q2. Âge", "Q3. Quartier", "Q4. Durée résidence Safi", 
        "Q5. Niveau d'études", "Q6. Situation Pro", "Q7. A fréquenté cinéma", "Q8. Périodes", "Q9. Fréquence", 
        "Q10. Accompagnants", "Q11. Types films", "Q11. Autre film", "Q12. Souvenir marquant", "Q13. Liste salles", 
        "Q14. Devenir salles", "Q15_1 TV", "Q15_2 Streaming", "Q15_3 Piratage", "Q15_4 Prix billet", 
        "Q15_5 Vétusté", "Q15_6 Qualité films", "Q15_7 Insécurité", "Q15_8 Prix terrain", 
        "Q15_9 Soutien public", "Q15_10 Changement loisirs", "Q16. Cause principale", "Q17_1 Souvenirs", 
        "Q17_2 Perte Safi", "Q17_3 État dégradé", "Q17_4 Attachement", "Q17_5 Identité", 
        "Q17_6 Jeunes ignorent", "Q18. Représentation", "Q19_1 Patrimoine", "Q19_2 Préserver", 
        "Q19_3 Dynamiser centre", "Q19_4 Créer emplois", "Q19_5 Fréquenter", "Q19_6 Associer habs", 
        "Q20. Usage souhaité", "Q20. Autre usage", "Q21. Soutien projet", "Q22. Vu média", 
        "Q23. Canaux", "Q24. Pages suivies", "Q25_1 Trop peu média", "Q25_2 Réseaux info", "Q25_3 Savoir plus", 
        "Q26. Suggestions", "Q27. Recontact", "Q27. Contact infos"
    ]
    
    # Write headers
    ws1.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write rows
    for r in responses:
        ws1.append([
            r['id'], r['submission_date'], r['q1_gender'], r['q2_age_group'], r['q3_neighborhood'], r['q4_residence_duration'],
            r['q5_education_level'], r['q6_profession'], r['q7_visited_cinema'], ", ".join(r['q8_periods']), r['q9_frequency'],
            ", ".join(r['q10_companions']), ", ".join(r['q11_movie_types']), r['q11_other'], r['q12_memory'], r['q13_text'],
            ", ".join(r['q14_what_became']), r['q15_1'], r['q15_2'], r['q15_3'], r['q15_4'],
            r['q15_5'], r['q15_6'], r['q15_7'], r['q15_8'], r['q15_9'], r['q15_10'], r['q16_main_cause'],
            r['q17_1'], r['q17_2'], r['q17_3'], r['q17_4'], r['q17_5'], r['q17_6'], r['q18_meaning'],
            r['q19_1'], r['q19_2'], r['q19_3'], r['q19_4'], r['q19_5'], r['q19_6'],
            ", ".join(r['q20_desired_usage']), r['q20_other'], ", ".join(r['q21_support_type']), r['q22_seen_content'],
            ", ".join(r['q23_channels']), r['q24_follow_pages'], r['q25_1'], r['q25_2'], r['q25_3'],
            r['q26_comments'], r['q27_recontact'], r['q27_contact_details']
        ])
        
    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        
    # Sheet 2: Cinema mentions
    ws2 = wb.create_sheet(title="Salles Citees (Q13)")
    headers_cm = ["N° Questionnaire", "Nom de la salle", "Quartier / Localisation", "État actuel"]
    ws2.append(headers_cm)
    for col_num, header in enumerate(headers_cm, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    for r in responses:
        for c in r.get('q13_table', []):
            ws2.append([r['id'], c.get('name'), c.get('location'), c.get('current_state')])
            
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 15), 50)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"enquete_safi_cinema_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

@app.route('/admin/api/export/pdf', methods=['POST'])
@login_required
def export_pdf():
    try:
        req_data = request.json or {}
        filters = req_data.get('filters', {})
        chart_images = req_data.get('charts', {}) # Dictionary of name: base64_png
        
        responses = get_all_responses(filters)
        stats = compile_full_stats(responses)
        
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=A4,
            leftMargin=0.5*inch, rightMargin=0.5*inch,
            topMargin=0.5*inch, bottomMargin=0.5*inch
        )
        
        styles = getSampleStyleSheet()
        
        # Setup custom styles for premium look
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=colors.HexColor('#1E3A8A'),
            alignment=1, # Center
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=11,
            textColor=colors.HexColor('#4B5563'),
            alignment=1,
            spaceAfter=25
        )
        h1_style = ParagraphStyle(
            'H1',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=14,
            textColor=colors.HexColor('#1E3A8A'),
            spaceBefore=15,
            spaceAfter=10,
            keepWithNext=True
        )
        text_style = ParagraphStyle(
            'Body',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=6,
            leading=14
        )
        bold_text_style = ParagraphStyle(
            'BodyBold',
            parent=text_style,
            fontName='Helvetica-Bold'
        )
        
        story = []
        
        # Header / Title Block
        story.append(Paragraph("ROYAUME DU MAROC - MASTER EN SIC", ParagraphStyle('SIC', parent=subtitle_style, fontSize=9, bold=True, spaceAfter=2)))
        story.append(Paragraph("Rapport d'Analyse Statistique de l'Enquête", title_style))
        story.append(Paragraph(f"Le patrimoine des salles de cinéma fermées de Safi : mémoire, perceptions et perspectives de valorisation", subtitle_style))
        
        # Metadata Block
        meta_data = [
            [Paragraph("<b>Date du rapport:</b>", text_style), Paragraph(datetime.now().strftime("%d/%m/%Y %H:%M"), text_style),
             Paragraph("<b>Échantillon analysé:</b>", text_style), Paragraph(f"{len(responses)} répondants", text_style)],
            [Paragraph("<b>Filtres appliqués:</b>", text_style), Paragraph(json.dumps(filters) if filters else "Aucun (échantillon global)", text_style),
             Paragraph("<b>Statut:</b>", text_style), Paragraph("Données officielles", text_style)]
        ]
        meta_table = Table(meta_data, colWidths=[1.5*inch, 2.2*inch, 1.5*inch, 2.0*inch])
        meta_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F3F4F6')),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Section A: Profil
        story.append(Paragraph("1. Profil de l'échantillon (Section A)", h1_style))
        
        # Display gender & age summary
        q1_label = "Sexe"
        q1_data = stats.get('q1', {})
        q2_label = "Tranche d'âge"
        q2_data = stats.get('q2', {})
        
        profile_table_data = [
            [Paragraph("<b>Question</b>", bold_text_style), Paragraph("<b>Modalité</b>", bold_text_style), Paragraph("<b>Effectif (N)</b>", bold_text_style), Paragraph("<b>Pourcentage (%)</b>", bold_text_style)]
        ]
        for val, info in q1_data.get('distribution', {}).items():
            profile_table_data.append([Paragraph("Q1. Genre", text_style), Paragraph(val, text_style), Paragraph(str(info['count']), text_style), Paragraph(f"{info['percentage']}%", text_style)])
        for val, info in q2_data.get('distribution', {}).items():
            profile_table_data.append([Paragraph("Q2. Tranche d'âge", text_style), Paragraph(val, text_style), Paragraph(str(info['count']), text_style), Paragraph(f"{info['percentage']}%", text_style)])
            
        profile_table = Table(profile_table_data, colWidths=[2.0*inch, 2.2*inch, 1.5*inch, 1.5*inch])
        profile_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        # Color header columns text white
        for cell in profile_table_data[0]:
            cell.style.textColor = colors.white
            
        story.append(profile_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Inject Chart 1 if exists
        if 'chart_profil' in chart_images:
            try:
                img_data = base64.b64decode(chart_images['chart_profil'].split(',')[1])
                img_file = io.BytesIO(img_data)
                story.append(RLImage(img_file, width=5.5*inch, height=2.5*inch))
                story.append(Spacer(1, 0.2*inch))
            except Exception as e:
                print(f"Error rendering profil chart in PDF: {e}")
                
        # Page Break
        story.append(PageBreak())
        
        # Section B & C: Mémoire & Fréquentation
        story.append(Paragraph("2. Mémoire et Fréquentation des salles (Section B & C)", h1_style))
        q7_data = stats.get('q7', {})
        visited_pct = q7_data.get('distribution', {}).get('Oui', {}).get('percentage', 0)
        story.append(Paragraph(f"Taux de fréquentation historique des salles de cinéma à Safi : <b>{visited_pct}%</b> des répondants déclarent avoir déjà fréquenté une ou plusieurs salles de cinéma dans la ville.", text_style))
        
        # Inject Chart 2 if exists (Memoire/Fréquence)
        if 'chart_visited' in chart_images:
            try:
                img_data = base64.b64decode(chart_images['chart_visited'].split(',')[1])
                img_file = io.BytesIO(img_data)
                story.append(RLImage(img_file, width=5.5*inch, height=2.8*inch))
                story.append(Spacer(1, 0.2*inch))
            except Exception as e:
                print(f"Error rendering visited chart in PDF: {e}")
                
        # Section D: Causes de Fermeture
        story.append(Paragraph("3. Causes de fermeture des salles (Section D)", h1_style))
        story.append(Paragraph("Échelles de Likert : moyennes et distributions des avis concernant les facteurs de fermeture.", text_style))
        
        q15_data = stats.get('q15', {}).get('items', {})
        likert_table_data = [
            [Paragraph("<b>Affirmation (Facteur de fermeture)</b>", bold_text_style), Paragraph("<b>Moyenne / 5</b>", bold_text_style), Paragraph("<b>Médiane</b>", bold_text_style)]
        ]
        for col, item in q15_data.items():
            likert_table_data.append([
                Paragraph(item['statement'], text_style),
                Paragraph(str(item['mean']), text_style),
                Paragraph(item['median'], text_style)
            ])
            
        likert_table = Table(likert_table_data, colWidths=[4.2*inch, 1.5*inch, 1.5*inch])
        likert_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        for cell in likert_table_data[0]:
            cell.style.textColor = colors.white
            
        story.append(likert_table)
        story.append(Spacer(1, 0.2*inch))
        
        if 'chart_closing' in chart_images:
            try:
                img_data = base64.b64decode(chart_images['chart_closing'].split(',')[1])
                img_file = io.BytesIO(img_data)
                story.append(RLImage(img_file, width=5.5*inch, height=3.0*inch))
                story.append(Spacer(1, 0.2*inch))
            except Exception as e:
                print(f"Error rendering closing chart in PDF: {e}")
                
        # Page Break
        story.append(PageBreak())
        
        # Section E & F: Attachement & Perspectives de valorisation
        story.append(Paragraph("4. Représentations, attachement et avenir (Section E & F)", h1_style))
        
        q17_data = stats.get('q17', {}).get('items', {})
        q19_data = stats.get('q19', {}).get('items', {})
        
        att_table_data = [
            [Paragraph("<b>Affirmation (Représentations & Perspectives)</b>", bold_text_style), Paragraph("<b>Moyenne / 5</b>", bold_text_style), Paragraph("<b>Médiane</b>", bold_text_style)]
        ]
        # Top 3 Q17 items
        for col in ['q17_2', 'q17_4', 'q17_5']:
            if col in q17_data:
                att_table_data.append([Paragraph(q17_data[col]['statement'], text_style), Paragraph(str(q17_data[col]['mean']), text_style), Paragraph(q17_data[col]['median'], text_style)])
        # Top 3 Q19 items
        for col in ['q19_1', 'q19_2', 'q19_3']:
            if col in q19_data:
                att_table_data.append([Paragraph(q19_data[col]['statement'], text_style), Paragraph(str(q19_data[col]['mean']), text_style), Paragraph(q19_data[col]['median'], text_style)])
                
        att_table = Table(att_table_data, colWidths=[4.2*inch, 1.5*inch, 1.5*inch])
        att_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        for cell in att_table_data[0]:
            cell.style.textColor = colors.white
        story.append(att_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Q20: Usage souhaité
        story.append(Paragraph("<b>Usages souhaités en cas de réhabilitation (Q20) :</b>", bold_text_style))
        q20_freq = stats.get('q20', {}).get('frequencies', {})
        usage_bullets = []
        for usage, info in list(q20_freq.items())[:5]: # Top 5
            usage_bullets.append(f"• {usage} ({info['percentage']}% des répondants)")
        story.append(Paragraph("<br/>".join(usage_bullets), text_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Word cloud / open responses keywords summary
        story.append(Paragraph("5. Analyse sémantique des réponses ouvertes", h1_style))
        story.append(Paragraph("<b>Mots-clés récurrents concernant le ressenti des citoyens (Q18 : En un mot, que représentent ces salles ?) :</b>", bold_text_style))
        q18_words = stats.get('q18', {}).get('word_frequencies', [])
        word_list = [f"<b>{w['text']}</b> ({w['value']} fois)" for w in q18_words[:15]]
        story.append(Paragraph(", ".join(word_list) if word_list else "Aucun verbatim significatif.", text_style))
        
        doc.build(story)
        pdf_buffer.seek(0)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"rapport_enquete_cinema_safi_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
    except Exception as e:
        print(f"Error generating PDF: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Erreur lors de la génération du PDF : {e}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
